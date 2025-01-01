import tempfile
from django.conf import settings
from django.shortcuts import render
from django.http import FileResponse
from PyPDF2 import PdfMerger
from PIL import Image
from io import BytesIO
import os
import pandas as pd
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter,landscape,A4
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
# from django.core.files import File
import pdfplumber
import pandas
import pdfkit
import tabula
import io
from docx2pdf import convert
import xlsxwriter

def convert_file(file_object, conversion_format):
    filename = file_object.name
    name, file_extension = os.path.splitext(filename)

    converted_file = None
    uploads_dir = os.path.join(settings.MEDIA_ROOT, 'uploads')

    # Ensure the uploads directory exists
    if not os.path.exists(uploads_dir):
        os.makedirs(uploads_dir)

    # Determine the conversion and save the file to the specified path
    if file_extension in ('.jpg', '.jpeg', '.jfi') and conversion_format == 'PDF':
        converted_file = upload_jpg(file_object, conversion_format)  # Your conversion function here
        file_extension = '.pdf'
    elif file_extension in ('.xlsx', '.xls') and conversion_format == 'PDF':
        converted_file = xlsx_to_pdf(file_object)  # Your conversion function here
        file_extension = '.pdf'
    elif file_extension in ('.pdf') and conversion_format == 'CSV':
        pass  # You can add your conversion logic here
    elif file_extension in ('.pdf') and conversion_format == 'XLS':
        converted_file = pdf_excel(file_object, name)  # Your conversion function here
        file_extension = '.xls'
    elif file_extension in ('.docx') and conversion_format == 'PDF':
        convert_file =  docx_pdf(file_object)
    else:
        print(f'File extension {file_extension} is not supported for format {conversion_format}')
        return None

    # Generate a unique filename for the converted file (to avoid overwriting)
    filename_cr = f"{name}_converted{file_extension}"
    file_path = os.path.join(uploads_dir, filename_cr)

    # Now save the converted file content to the specified path
    with open(file_path, 'wb') as f:
        f.write(converted_file) 

    return file_path


def convert_FILE(file_object, conversion_format):
    filename = file_object.name
    # filename = 'another.jpg'
    name, file_extension = os.path.splitext(filename)
    print(name)

    converted_file = None

    if file_extension in ('.jpg', '.jpeg', '.jfi') and conversion_format == 'PDF':
        converted_file = upload_jpg(file_object, conversion_format)
    elif file_extension in ('.xlsx', '.xls') and conversion_format == 'PDF':
        converted_file = xlsx_pdf(file_object)
    elif file_extension in ('.pdf') and conversion_format == 'CSV':
        pass
        # converted_file = pdf_to_csv(file_object)
    elif file_extension in ('.pdf') and conversion_format == 'XLS':
        converted_file = pdf_excel(file_object, name)
    else:
        print(f'File extension {file_extension} is not supported for format {conversion_format}')

    return converted_file

def pdf_excel(file, name):
    excel_file = f'{name}.xls'
    with pdfplumber.open(file) as pdf:
        all_tables = []
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if table:
                    df = pd.DataFrame(table)
                    all_tables.append(df)

        if not tables:
            all_tables.append(pd.DataFrame([["No tables found"]]))
        
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            for idx, df in enumerate(all_tables):
                df.to_excel(writer, sheet_name=f'Sheet{idx+1}', index=False)

    return excel_file

def upload_jpg(file, format):
    """
    This function converts the `.jpg` and `.jpeg` files into one `.pdf` file.
    Returns the PDF content as bytes.
    """
    pdf_bytes = None
    merger = PdfMerger()

    filename = file.name
    name, _ = os.path.splitext(filename)

    try:
        # Open the image and verify it
        image = Image.open(file)
        print(f"Valid image format: {image.format}")

        # Handle transparency (RGBA, LA, etc.)
        if image.mode in ("RGBA", "LA") or (image.mode == "P" and "transparency" in image.info):
            image.load()
            background = Image.new("RGB", image.size, (255, 255, 255))  # White background
            background.paste(image, mask=image.split()[3])  # Paste with transparency mask
            image = background

        # Save the image as a PDF in memory using BytesIO (in-memory file-like object)
        pdf_bytes_io = io.BytesIO()  # Create an in-memory byte stream

        # Save image as PDF into the byte stream
        image.save(pdf_bytes_io, 'PDF', encoding='latin-1')
        pdf_bytes_io.seek(0)  # Rewind the in-memory byte stream to the beginning

        # Append the in-memory PDF file to the merger
        merger.append(pdf_bytes_io)

        # Return the PDF as bytes
        pdf_bytes = pdf_bytes_io.getvalue()

    except Exception as e:
        print(f"Error: {e}")

    return pdf_bytes  # Return the PDF content as bytes


def xlxs_pdff(file):
    # workbook = xlsxwriter.workbook(file)
    pass

    


# docx2pdf is not implemented for linux 
def docx_pdf(file):
    # Ensure the temporary file has the correct '.docx' suffix
    with tempfile.NamedTemporaryFile(delete=False, suffix='.docx') as temp_file:
        # Write the content of the uploaded file to the temporary file
        for chunk in file.chunks():
            temp_file.write(chunk)
        
        # The path to the temporary file
        temp_file_path = temp_file.name
        print(f'Temporary DOCX file saved at: {temp_file_path}')

        try:
            # Generate the output PDF file path by replacing '.docx' with '.pdf'
            converted_pdf_path = temp_file_path.replace('.docx', '.pdf')
            print(f'Converted PDF will be saved at: {converted_pdf_path}')
            
            # Perform the conversion and save to the converted PDF path
            convert(temp_file_path, converted_pdf_path)
            print(f'PDF saved at: {converted_pdf_path}')
        except Exception as e:
            print(f'Error during conversion: {e}')
            return None
    
    # Return the path to the converted PDF
    return converted_pdf_path


def upload_jpgss(file, format):
    """
    This function converts the `.jpg` and `.jpeg` files into one `.pdf` file.
    """
    pdfFile = None
    merger = PdfMerger()

    filename = file.name
    name, _ = os.path.splitext(filename)

    try:
        # Open the image and verify it
        image = Image.open(file)
        print(f"Valid image format: {image.format}")

        # Handle transparency (RGBA, LA, etc.)
        if image.mode in ("RGBA", "LA") or (image.mode == "P" and "transparency" in image.info):
            image.load()
            background = Image.new("RGB", image.size, (255, 255, 255))  # White background
            background.paste(image, mask=image.split()[3])  # Paste with transparency mask
            image = background

        # Save the image as a PDF in memory using BytesIO (in-memory file-like object)
        pdf_bytes_io = io.BytesIO()  # Create an in-memory byte stream

        # Save image as PDF into the byte stream
        image.save(pdf_bytes_io, 'PDF', encoding='latin-1')
        pdf_bytes_io.seek(0)  # Rewind the in-memory byte stream to the beginning

        # Append the in-memory PDF file to the merger
        merger.append(pdf_bytes_io)

    except Exception as e:
        print(f"Error: {e}")

    # Save the merged PDF file
    format = format.lower()
    pdfFile = f"{name}.{format}"

    # Write the final merged PDF to a file
    merger.write(pdfFile)
    merger.close()
    print(f"PDF saved as {pdfFile}")

    return pdfFile  # Return the path to the final merged PDF


# no wbfilkmf library
def xlsx_to_pdf(file):
    filename = file.name
    name,_ = os.path.splitext(filename)
    output = f'{name}.pdf'
    
    df = pd.read_excel(file)
    df.to_html('file.html')
    pdfkit.from_file('file.html', output)
    
    with open(output, 'rb') as f:
        pdf_file = f.read()
    print(type(pdf_file))
    return pdf_file


def pdf_to_csv(file):
    filename = file.name
    name,_ = os.path.splitext(filename)
    output = f'{name}.csv'

    csvFile = tabula.convert_into(file, output)
    return csvFile

# converts to str not byte
def xlsx_pdf(file):
    """
        This function converts all excel files `.xlsx`, `.xlx` into `pdf`
    """
    pdfFile = "output.pdf"
    # for file in files:
    print(file, '....................................................')

    # load excel file
    workbook = load_workbook(file)

    # get the current sheet
    worksheet = workbook.active

    # Read the number of rows and columns in the current sheet
    max_row = worksheet.max_row
    max_column = worksheet.max_column

    # set PDF canvas
    c = canvas.Canvas(pdfFile, pagesize=landscape(A4))

    # set margins
    top_margin = 3*inch
    left_margin = 0.75*inch 
    bottom_margin = 0.75*inch
    right_margin = 0.75*inch


    cell_width = (11*inch - left_margin - right_margin) / max_column
    cell_height = (8.5*inch - top_margin - bottom_margin) / max_row

    for row in range(1, max_row+1):
        for column in range(1, max_column+1):
            cell = worksheet.cell(row=row, column=column)
            text=str(cell.value)
            x = left_margin + (column -1) * cell_width
            # y = 11*inch - (top_margin + max_row * cell_height)
            y = 11*inch - (top_margin + row * cell_height)
            c.drawString(x, y, text)

    # save PDF and close canvas
    c.save()
    print('Done conversion of .xls to pdf file')
    return pdfFile






def xlsx_to0000_pdf(file):
    filename = file.name
    name,_ = os.path.splitext(filename)
    output = f'{name}.pdf'
    
    df = pd.read_excel(file)
    df.to_html('file.html')
    pdFile = pdfkit.from_file('file.html', output)
    print(type(pdFile))
    return pdFile





# def xlsx_2_pdf(file):
#     # if request.method == 'POST':
#     #     xlsxFiles = request.FILES.getlist('file')

#     #     if not xlsxFiles:
#     #         return render(request, 'converter/uploadfile.html', {'error': 'Please select a file to upload'})

#     #     for xfile in xlsxFiles:

#     # Read the data from the Excel file
#     data = pd.read_excel(file)

#     # Create a PDF document
#     pdfFile = SimpleDocTemplate('excelfile.pdf', pagesize=letter)

#     # Create a table with the data
#     tabel = Table(data.values)

#     # Add some style to the table
#     tabel.setStyle(TableStyle([
#         ('BACKGROUND', (0, 0), (-1, 0), '#d0d0d0'),
#         ('TEXTCOLOR', (0, 0), (-1, 0), '#000000'),
#         ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#         ('FONTSIZE', (0, 0), (-1, 0), 14),
#         ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
#         ('BACKGROUND', (0, 1), (-1, -1), '#f5f5f5'),
#     ]))

#     # Add tabel to the PDF document
#     pdfFile.build([tabel])

#     return pdfFile